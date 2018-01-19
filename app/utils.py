import argparse
import csv
import os
import logging

from openpyxl import Workbook, load_workbook
from slacker import Slacker

XLSX_FILE_NAME = "results.xlsx"
CSV_FILE_NAME = "results.csv"
DIR_NAME = "tests\_test_results"
logger = logging.getLogger(__name__)


class Utils(object):
    @staticmethod
    def send_msg_to_slack(_channel=None, _msg=''):
        slack_token = os.getenv("SLACK_TOKEN")
        if not slack_token:
            return "SLACK_TOKEN not defined"

        slack = Slacker(slack_token)
        slack.chat.post_message(_channel, str(_msg))
        logger.info("Sent message to slack channel '{}':\n '{}' ".format(_channel, _msg))

    @staticmethod
    def is_string_in_csv(_string, file_name, dir_name=DIR_NAME):
        filepath = os.path.join(dir_name, file_name)
        if not os.path.exists(dir_name):
            os.makedirs(dir_name)

        if not os.path.isfile(filepath):
            open(filepath, 'a').close()

        with open(filepath, mode='r', encoding="utf8") as f:
            reader = csv.reader(f)
            for row in reader:
                if _string in row:
                    return True

    @staticmethod
    def create_parser():
        _parser = argparse.ArgumentParser(
            description='The script saves jobs info from freelancer.com to csv and xlsx-file.')
        _parser.add_argument('-k', '--keyword', help='keyword for searching', default='KEYWORD')
        _parser.add_argument('-p', '--pages', help='number of pages for scraping', default=1)
        return _parser

    # TODO: Merge write methods
    # TODO: DO not add existing title to file 
    @staticmethod
    def write_csv(data, file_name=CSV_FILE_NAME, dir_name=DIR_NAME):
        if not data:
            logger.info("Nothing write to csv-file")
            return None

        filepath = os.path.join(dir_name, file_name)
        if not os.path.exists(dir_name):
            os.makedirs(dir_name)

        with open(filepath, 'a+', newline='', encoding='utf-8') as csvfile:
            filewriter = csv.writer(csvfile, delimiter=',', quoting=csv.QUOTE_MINIMAL)
            # Add header if not exists
            if not Utils.is_string_in_csv(data[0].keys(), file_name):
                filewriter.writerow(data[0].keys())
            # Add row to file
            for row in data:
                if not Utils.is_string_in_csv(row.get('title'), file_name):
                    filewriter.writerow([value for value in row.values()])
            logger.info("File {} saved".format(filepath))

    @staticmethod
    def write_xlsx(data, sheet_title='NewSheet', file_name=XLSX_FILE_NAME, dir_name=DIR_NAME):
        # TODO: Do not create new sheet everytime 
        if not data:
            logger.info("Nothing write to csv-file")
            return None

        filepath = os.path.join(dir_name, file_name)
        if not os.path.exists(dir_name):
            os.makedirs(dir_name)

        try:
            wb = load_workbook(filepath)
        except FileNotFoundError:
            wb = Workbook()
        ws = wb.create_sheet(sheet_title)
        ws.append(list(data[0].keys()))
        if data:
            for item_info in data:
                ws.append([value for value in item_info.values()])
        wb.save(filepath)
        logger.info("File {} saved".format(filepath))

    @staticmethod
    def send_results_to_slack(channel='jobs', results=None, file_name=None):
        msg = ""
        for result in results:
            # Don't add duplicates
            if Utils.is_string_in_csv(result.get("title"), file_name):
                break

            for k, v in result.items():
                if k == "title":
                    msg += "\n *{k}: {v}*".format(k=k, v=v)
                else:
                    msg += "\n {k}: {v}".format(k=k, v=v)
        if msg:
            Utils.send_msg_to_slack(channel, msg)
        else:
            logger.info("Nothig sent to slack. Msg is empty")
